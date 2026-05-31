"""BNR interactive-database (Baza de date interactivă) historical ROBOR import.

One-time historical backfill source for ROBOR. The live daily scraper
(``robor.py`` / curs-valutar-bnr.ro) only exposes a ~5-month rolling window, so
ROBOR before that was missing from Themis. The BNR BDI export
("ROBID - ROBOR - serii zilnice", available since August 1995) fills the gap.

Export layout (verified 2026-05-31):
  - Rows 1-4: metadata ("Nume clasa statistica:", "Nota:", ...).
  - A header row whose first cell is "Data".
  - Header columns interleave ROBID <tenor> ... then ROBOR <tenor> ...
    ("ROBOR overnight", "ROBOR 1 săptămână", "ROBOR 1 lună", "ROBOR 3 luni",
     "ROBOR 6 luni", "ROBOR 9 luni", "ROBOR 12 luni", plus "ROBOR tomorrow next").
  - Two sub-header rows ("(% p.a.)" units; "BBZ_*" codes) with an empty date cell.
  - Data rows: date (YYYY-MM-DD string, or a datetime), then floats; missing
    values are "-".

We import ROBOR ONLY (ROBID is out of scope for loan accrual) and only the
tenors consumers expect — ON, 1W, 1M, 3M, 6M, 12M — skipping tomorrow-next and
9M, matching ``robor.py``'s tenor set so the historical rows are
indistinguishable from the daily scraper's. Idempotent via
``store_interest_rates`` (INSERT OR IGNORE on (date, rate_type, tenor)).
"""
from __future__ import annotations

import datetime
import logging
import unicodedata
from typing import Iterable

from app.services.rates.robor import ParsedInterestRate

logger = logging.getLogger(__name__)


def _strip_diacritics(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def _norm(value: object) -> str:
    """Lowercase + diacritic-stripped + whitespace-collapsed cell text."""
    if value is None:
        return ""
    return " ".join(_strip_diacritics(str(value)).strip().lower().split())


# Diacritic-stripped, normalized ROBOR header -> tenor. Tomorrow-next and 9M
# are intentionally absent (skipped) to match robor.py's consumer tenor set.
_BDI_ROBOR_HEADER_TO_TENOR = {
    "robor overnight": "ON",
    "robor 1 saptamana": "1W",
    "robor 1 luna": "1M",
    "robor 3 luni": "3M",
    "robor 6 luni": "6M",
    "robor 12 luni": "12M",
}


def _parse_date_cell(value: object) -> str | None:
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        try:
            return datetime.date.fromisoformat(s).isoformat()
        except ValueError:
            return None
    return None


def _parse_rate_cell(value: object) -> float | None:
    if isinstance(value, bool):  # guard: bool is an int subclass
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().rstrip("%").strip().replace(",", ".")
        if not s or s == "-":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def parse_bnr_bdi_rows(rows: Iterable[tuple]) -> list[ParsedInterestRate]:
    """Parse BNR BDI worksheet rows into ROBOR ParsedInterestRate values.

    Locates the header row (first cell == "Data"), maps the ROBOR columns we
    want, then walks the data rows. Rows without a parseable date (metadata,
    units/codes sub-headers, blanks) are skipped; "-"/blank rate cells are
    skipped. Pure — no I/O — so it's unit-testable without an .xlsx.
    """
    materialized = [tuple(r) if r is not None else () for r in rows]

    header_idx: int | None = None
    col_to_tenor: dict[int, str] = {}
    for i, row in enumerate(materialized):
        if not row or _norm(row[0]) != "data":
            continue
        for ci, cell in enumerate(row):
            tenor = _BDI_ROBOR_HEADER_TO_TENOR.get(_norm(cell))
            if tenor is not None:
                col_to_tenor[ci] = tenor
        header_idx = i
        break

    if header_idx is None or not col_to_tenor:
        return []

    out: list[ParsedInterestRate] = []
    for row in materialized[header_idx + 1:]:
        if not row:
            continue
        date = _parse_date_cell(row[0])
        if not date:
            continue
        for ci, tenor in col_to_tenor.items():
            if ci >= len(row):
                continue
            rate = _parse_rate_cell(row[ci])
            if rate is None:
                continue
            out.append(
                ParsedInterestRate(date=date, rate_type="ROBOR", tenor=tenor, rate=rate)
            )
    return out


def read_bnr_bdi_xlsx(path: str) -> list[ParsedInterestRate]:
    """Read a BNR BDI ROBID-ROBOR .xlsx export and return ROBOR rows.

    openpyxl is imported lazily so the app runtime (and the prod container,
    which has no openpyxl) never depends on it — only the local seed-build
    step does. The built seed is loaded in prod via ``read_seed_csv_gz``.
    """
    import openpyxl  # lazy: import-only dependency for the local seed build

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return parse_bnr_bdi_rows(ws.iter_rows(values_only=True))
    finally:
        wb.close()


# Header for the gzipped-CSV seed that build_robor_seed.py writes and
# load_robor_seed.py reads. Stdlib-only round-trip (gzip + csv) so it works in
# the prod container without openpyxl.
SEED_CSV_FIELDS = ("date", "rate_type", "tenor", "rate")


def read_seed_csv_gz(path: str) -> list[ParsedInterestRate]:
    """Read a gzipped-CSV ROBOR seed (date,rate_type,tenor,rate) into rows.

    Standard-library only — this is the path used in production. The seed is
    produced from the BNR .xlsx export by scripts/build_robor_seed.py.
    """
    import csv
    import gzip

    out: list[ParsedInterestRate] = []
    with gzip.open(path, "rt", newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            rate = _parse_rate_cell(row.get("rate"))
            date = _parse_date_cell(row.get("date"))
            if date is None or rate is None:
                continue
            out.append(
                ParsedInterestRate(
                    date=date,
                    rate_type=(row.get("rate_type") or "ROBOR"),
                    tenor=(row.get("tenor") or ""),
                    rate=rate,
                )
            )
    return out
